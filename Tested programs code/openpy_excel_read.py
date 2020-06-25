import openpyxl

wk= openpyxl.load_workbook("C:\DRIVE DATA\TestData.xlsx")
sh= wk["Sheet1"]
sh1=wk.active.title
print(sh1)
rows = sh.max_row
cols = sh.max_column
print("max rows" +str(rows))
print("max columns" +str(cols))

#reading data in 2 different ways



for r in sh['A1':'C8']:
    for c in r:
        print(c.value)

""""
for i in range(1,rows+1):
    for j in range(1,cols+1):
        wc=sh.cell(i,j)
        print(wc.value)

""""





#1st approach to read cell value
#print(sh['A1'].value)
#print(sh['B1'].value)

#2nd approach to read cell value
#c1=sh.cell(1,1)
#c1=sh.cell(row=1, column=1)
#print(c1.value)
#print(c1.row)
#print(c1.column)

#c1=sh.cell(1,2)
#print(c1.value)



